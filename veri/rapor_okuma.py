"""e-Okul rapor ayrıştırıcıları.

İki rapor okunur:
  * OOK12001R010 — Sorumluluk sınavına girecek öğrenci listesi
  * OOK01001R1   — Kurum personel listesi

Bu modül veritabanına dokunmaz; dosyadan kayıt üretir ve doğrular. Böylece
gerçek rapor düzenine karşı bağımsız test edilebilir.
"""

from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path

from cekirdek.metin import esitle, sadelestir
from cekirdek.modeller import SorumlulukKaydi, mudur_mu


# "... Anadolu Lisesi - 9. Sınıf / A Şubesi" biçimindeki grup başlığı.
SUBE_DESENI = re.compile(r"^.+-\s*(\d+)\.\s*Sınıf\s*/\s*(.+?)\s*Şubesi", re.IGNORECASE)

# OOK12001R010'da sabit sütun yerleşimi (0 tabanlı).
SUT_OGRENCI_NO = 1
SUT_AD_SOYAD = 2
SUT_SINIF = 8
SUT_DERS = 9


class RaporHatasi(ValueError):
    """Rapor okunamadığında kullanıcıya gösterilecek hata."""


def dosya_ozeti(yol: Path) -> str:
    return hashlib.sha256(Path(yol).read_bytes()).hexdigest()


def _sayiyi_temizle(deger: str) -> str:
    """Excel sayısal hücrelerinin '123.0' biçimini '123' yapar."""
    return deger[:-2] if deger.endswith(".0") else deger


# ------------------------------------------------------------- tablo okuma

def _xlsx_satirlari(yol: Path) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as hata:  # pragma: no cover - kurulum hatası
        raise RaporHatasi(".xlsx raporu için openpyxl paketi kurulmalıdır.") from hata
    kitap = load_workbook(yol, read_only=True, data_only=True)
    try:
        return [list(satir) for satir in kitap.active.iter_rows(values_only=True)]
    finally:
        kitap.close()


def _xls_satirlari(yol: Path) -> list[list[object]]:
    try:
        import xlrd
    except ImportError as hata:  # pragma: no cover - kurulum hatası
        raise RaporHatasi(".xls raporu için xlrd paketi kurulmalıdır.") from hata
    kitap = xlrd.open_workbook(str(yol))
    sayfa = kitap.sheet_by_index(0)
    return [sayfa.row_values(sira) for sira in range(sayfa.nrows)]


def _csv_satirlari(yol: Path) -> list[list[object]]:
    with Path(yol).open("r", encoding="utf-8-sig", newline="") as akim:
        return [list(satir) for satir in csv.reader(akim)]


def tablo_oku(yol: Path) -> list[list[object]]:
    """Desteklenen biçimlerden ham satır listesi üretir."""
    yol = Path(yol)
    if not yol.exists():
        raise RaporHatasi(f"Dosya bulunamadı: {yol.name}")
    okuyucular = {".xlsx": _xlsx_satirlari, ".xls": _xls_satirlari, ".csv": _csv_satirlari}
    okuyucu = okuyucular.get(yol.suffix.lower())
    if okuyucu is None:
        raise RaporHatasi("Yalnız .xls, .xlsx ve .csv dosyaları desteklenir.")
    return okuyucu(yol)


# ================================================ OOK12001R010 — sorumluluk

@dataclass(frozen=True)
class SorumlulukRaporu:
    kayitlar: tuple[SorumlulukKaydi, ...]
    dosya_ozeti: str
    subeler: tuple[str, ...]

    @property
    def ogrenci_sayisi(self) -> int:
        return len({(k.okul_no, k.sube) for k in self.kayitlar})

    @property
    def sube_sayisi(self) -> int:
        return len(self.subeler)

    @property
    def ders_duzey_sayisi(self) -> int:
        return len({(k.ders_adi, k.sinif_duzeyi) for k in self.kayitlar})

    def ogrenci_yukleri(self) -> dict[str, int]:
        """Öğrenci anahtarı -> sorumlu olduğu ders sayısı."""
        yukler: dict[str, int] = {}
        for kayit in self.kayitlar:
            yukler[kayit.ogrenci_anahtari] = yukler.get(kayit.ogrenci_anahtari, 0) + 1
        return yukler

    @property
    def azami_ogrenci_yuku(self) -> int:
        return max(self.ogrenci_yukleri().values(), default=0)


def sorumluluk_raporu_coz(satirlar: list[list[object]], dosya_ozeti: str) -> SorumlulukRaporu:
    """OOK12001R010 satırlarından sorumluluk kayıtları üretir.

    Rapor şube başlıklarıyla bölünmüştür; öğrenci no ve ad yalnız ilk satırda
    yazılıp izleyen ders satırlarında boş bırakılır, bu yüzden son görülen
    değerler taşınır.
    """
    kayitlar: list[SorumlulukKaydi] = []
    subeler: list[str] = []
    etkin_sube = ""
    okul_no = ""
    ad_soyad = ""

    for satir in satirlar:
        hucreler = list(satir) + [None] * max(0, 10 - len(satir))
        basligi = sadelestir(hucreler[0])
        eslesme = SUBE_DESENI.search(basligi)
        if eslesme:
            etkin_sube = f"{eslesme.group(1)}/{eslesme.group(2).strip()}"
            if etkin_sube not in subeler:
                subeler.append(etkin_sube)
            okul_no = ad_soyad = ""
            continue

        no_hucresi = sadelestir(hucreler[SUT_OGRENCI_NO])
        ad_hucresi = sadelestir(hucreler[SUT_AD_SOYAD])
        sinif_hucresi = sadelestir(hucreler[SUT_SINIF])
        ders_hucresi = sadelestir(hucreler[SUT_DERS])

        # Tablo başlığı satırları
        if no_hucresi == "Öğrenci No" or sinif_hucresi == "Sınıfı" or ders_hucresi == "Dersi":
            continue
        if no_hucresi:
            okul_no = _sayiyi_temizle(no_hucresi)
        if ad_hucresi:
            ad_soyad = ad_hucresi
        if not (etkin_sube and okul_no and ad_soyad and sinif_hucresi and ders_hucresi):
            continue
        try:
            duzey = int(float(sinif_hucresi))
        except ValueError:
            continue
        kayitlar.append(SorumlulukKaydi(okul_no, ad_soyad, etkin_sube, duzey, ders_hucresi))

    if not kayitlar:
        raise RaporHatasi(
            "Dosyada sorumluluk kaydı bulunamadı. e-Okul'dan OOK12001R010 raporunu "
            "biçimlendirmeden dışa aktardığınızdan emin olun."
        )
    return SorumlulukRaporu(tuple(kayitlar), dosya_ozeti, tuple(subeler))


def sorumluluk_raporu_oku(yol: Path) -> SorumlulukRaporu:
    return sorumluluk_raporu_coz(tablo_oku(yol), dosya_ozeti(yol))


# ==================================================== OOK01001R1 — personel

@dataclass(frozen=True)
class PersonelKaydi:
    satir_no: int
    ad: str
    unvan: str
    kadro_durumu: str
    brans: str
    personel_tipi: str
    kurum_sicil_no: str = ""


@dataclass(frozen=True)
class PersonelRaporu:
    kayitlar: tuple[PersonelKaydi, ...]
    dosya_ozeti: str

    @property
    def yonetici_sayisi(self) -> int:
        return sum(1 for k in self.kayitlar if "müdür" in esitle(k.unvan))

    @property
    def ogretmen_sayisi(self) -> int:
        return len(self.kayitlar) - self.yonetici_sayisi

    def branslar(self) -> tuple[str, ...]:
        gorulen: dict[str, str] = {}
        for kayit in self.kayitlar:
            gorulen.setdefault(esitle(kayit.brans), kayit.brans)
        return tuple(gorulen.values())


_ARANAN_BASLIKLAR = {
    "adı soyadı": "ad",
    "görevi": "unvan",
    "kadro durumu": "kadro",
    "branşı": "brans",
}
_ISTEGE_BAGLI_BASLIKLAR = {"kurum sicil no": "sicil", "t.c. kimlik no": None}


def _personel_tipi(kadro: str, unvan: str) -> str:
    anahtar = esitle(kadro)
    if "sözleşmeli" in anahtar:
        return "sozlesmeli"
    if "ücretli" in anahtar:
        return "ucretli"
    if "kadrolu" in anahtar:
        return "kadrolu"
    return "yonetici" if mudur_mu(unvan) or "müdür" in esitle(unvan) else "diger"


def personel_raporu_coz(satirlar: list[list[object]], dosya_ozeti: str) -> PersonelRaporu:
    """OOK01001R1 satırlarından personel kayıtları üretir.

    Sütun sırası raporda değişebildiği için başlık satırı aranarak eşleme
    kurulur; T.C. kimlik numarası sütunu varsa okunmaz ve saklanmaz.
    """
    baslik_sirasi = -1
    kolonlar: dict[str, int] = {}
    for sira, satir in enumerate(satirlar[:20]):
        bulunan = {esitle(str(deger)): indis for indis, deger in enumerate(satir)}
        if all(baslik in bulunan for baslik in _ARANAN_BASLIKLAR):
            baslik_sirasi = sira
            kolonlar = {hedef: bulunan[kaynak] for kaynak, hedef in _ARANAN_BASLIKLAR.items()}
            for kaynak, hedef in _ISTEGE_BAGLI_BASLIKLAR.items():
                if hedef and kaynak in bulunan:
                    kolonlar[hedef] = bulunan[kaynak]
            break
    if baslik_sirasi < 0:
        raise RaporHatasi(
            "e-Okul Personel Listesi başlıkları bulunamadı "
            "(ADI SOYADI, GÖREVİ, KADRO DURUMU, BRANŞI)."
        )

    def hucre(satir: list[object], ad: str) -> str:
        indis = kolonlar.get(ad, -1)
        return sadelestir(satir[indis]) if 0 <= indis < len(satir) else ""

    kayitlar: list[PersonelKaydi] = []
    gorulen: dict[str, int] = {}
    for sira, satir in enumerate(satirlar[baslik_sirasi + 1:], baslik_sirasi + 2):
        ad = hucre(satir, "ad")
        if not ad or esitle(ad).startswith("toplam personel sayısı"):
            continue
        unvan, kadro, brans = hucre(satir, "unvan"), hucre(satir, "kadro"), hucre(satir, "brans")
        if not (unvan and kadro and brans):
            raise RaporHatasi(f"{sira}. satırda zorunlu personel alanı eksik (görev/kadro/branş).")
        anahtar = esitle(ad)
        if anahtar in gorulen:
            raise RaporHatasi(
                f"{sira}. satırdaki personel {gorulen[anahtar]}. satırda da geçiyor. "
                "Aynı adlı iki personel varsa raporda kurum sicil numarası bulunmalıdır."
            )
        gorulen[anahtar] = sira
        kayitlar.append(PersonelKaydi(
            sira, ad, unvan, kadro, brans,
            _personel_tipi(kadro, unvan), _sayiyi_temizle(hucre(satir, "sicil")),
        ))

    if not kayitlar:
        raise RaporHatasi("Raporda içe alınabilir personel bulunamadı.")
    return PersonelRaporu(tuple(kayitlar), dosya_ozeti)


def personel_raporu_oku(yol: Path) -> PersonelRaporu:
    yol = Path(yol)
    if yol.suffix.lower() not in {".xls", ".xlsx"}:
        raise RaporHatasi("Personel raporu .xls veya .xlsx olmalıdır.")
    return personel_raporu_coz(tablo_oku(yol), dosya_ozeti(yol))
